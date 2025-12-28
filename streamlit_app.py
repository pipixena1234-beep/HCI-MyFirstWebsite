import streamlit as st
import pandas as pd
from fpdf import FPDF
from io import BytesIO
import zipfile
import json
import base64
import requests
import altair as alt
from datetime import datetime
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
from google.oauth2 import service_account
import openpyxl
from openpyxl.styles import Font
import base64

def add_custom_style(logo_path):
    # Base64 encoding for the logo
    try:
        with open(logo_path, "rb") as f:
            data = f.read()
            encoded = base64.b64encode(data).decode()
    except FileNotFoundError:
        encoded = ""

    st.markdown(
        f"""
        <style>
        /* 1. Main Backgrounds */
        .stApp {{ background-color: #FFDE59; }}
        
        /* 2. Sidebar Background - Purple */
        [data-testid="stSidebar"] {{ 
            background-color: #602b7b !important; 
        }}

        /* 3. Inject Logo to Top Left of Sidebar */
        [data-testid="stSidebarContent"]::before {{
            content: "";
            display: block;
            margin: 20px auto 10px 20px; /* Top, Right, Bottom, Left */
            width: 80px;  /* Adjust size as needed */
            height: 80px;
            background-image: url("data:image/png;base64,{encoded}");
            background-size: contain;
            background-repeat: no-repeat;
        }}

        /* 4. Light Green Metric Cards (FIXED SYNTAX) */
        [data-testid="stMetric"] {{
            background-color: #ffbc00 !important;
            padding: 15px !important;
            border-radius: 10px !important;
            border: 2px solid #000000 !important;
        }}

        /* 4. Global Font: Bold and Black */
        /* This targets almost all text elements in the app */
        .stApp, .stApp p, .stApp label, .stApp span, .stApp div {{
            color: #000000 !important;
            font-weight: bold !important;
        }}

        /* 5. Buttons - Radish Red (Kept from your previous code) */
        div.stButton > button, div.stDownloadButton > button {{
            background-color: #D0312D !important;
            color: white !important; /* Button text stays white for readability */
            border: none !important;
            border-radius: 8px !important;
            transition: 0.3s;
            font-weight: bold !important;
        }}
        
        div.stButton > button:hover, div.stDownloadButton > button:hover {{
            background-color: #A02623 !important;
            transform: scale(1.02);
        }}

        /* 6. Radish Red Checkboxes */
        div[data-testid="stCheckbox"] span[role="checkbox"] {{
            background-color: #D0312D !important;
            border-color: #D0312D !important;
        }}

        /* 7. Clean Transparent Containers for Charts */
        .vega-embed {{
            background-color: transparent !important;
        }}
        /* Ensure Sidebar text remains readable (if black is too dark on purple) */
        /* If you want sidebar text white, change this to white */
        [data-testid="stSidebar"] p, [data-testid="stSidebar"] label {{
            color: #000000 !important; 
        }}

       /* 4. FORCE TOP SKILL (Success) Background - Light Green */
        div[data-testid="stNotification"]:has(svg[title="Success"]) {{
            background-color: #90EE90 !important;
            color: black !important;
            border: 2px solid black !important;
            border-radius: 10px !important;
        }}
        
        /* 5. FORCE FOCUS AREA (Warning) Background - Radish Red */
        div[data-testid="stNotification"]:has(svg[title="Warning"]) {{
            background-color: #D0312D !important;
            color: white !important;
            border: 2px solid black !important;
            border-radius: 10px !important;
        }}
        
        /* Fix text color inside Focus Area specifically for white contrast */
        div[data-testid="stNotification"]:has(svg[title="Warning"]) p,
        div[data-testid="stNotification"]:has(svg[title="Warning"]) div {{
            color: white !important;
        }}
        </style>
        """,
        unsafe_allow_html=True
    )

# Call the function
add_custom_style("logo.png")

# GLOBAL CONFIGURATION (Fixes NameErrors)
month_order = ["Jan", "Feb", "March", "Apr", "May", "June", "July", "August", "Sept", "Oct", "Nov", "Dec"]
grade_order = ["A", "B", "C", "D", "F"]
grade_colors = ['#2ecc71', '#3498db', '#f1c40f', '#e67e22', '#e74c3c']
skills = ["Logic", "UI", "Animation", "Teamwork"] # Ensure this matches your Excel columns

# =====================================
# 1. Page Configuration
# =====================================
st.set_page_config(page_title="Student Progress System", layout="wide")
st.title("📊 Student Progress Report System")

# =====================================
# 2. GitHub Automation Sidebar
# =====================================
def push_schedule_to_github(new_datetime_str):
    repo_name = "pipixena1234-beep/HCI-MyFirstWebsite" 
    file_path = "schedule.json"
    try:
        token = st.secrets["GITHUB_TOKEN"]
        url = f"https://api.github.com/repos/{repo_name}/contents/{file_path}"
        headers = {"Authorization": f"token {token}", "Accept": "application/vnd.github.v3+json"}
        get_res = requests.get(url, headers=headers)
        sha = get_res.json().get("sha") if get_res.status_code == 200 else None
        content_dict = {"target_datetime": new_datetime_str}
        encoded_content = base64.b64encode(json.dumps(content_dict, indent=4).encode()).decode()
        payload = {"message": f"Update schedule to {new_datetime_str}", "content": encoded_content}
        if sha: payload["sha"] = sha
        return requests.put(url, json=payload, headers=headers).status_code
    except Exception as e: return str(e)

st.sidebar.subheader("⏰ Automation Schedule")
date_pick = st.sidebar.date_input("Target Date", datetime.now())
time_pick = st.sidebar.time_input("Target Time", datetime.now())
target_str = f"{date_pick.strftime('%Y-%m-%d')} {time_pick.strftime('%H:%M')}"

if st.sidebar.button("Update GitHub Schedule"):
    with st.spinner("Pushing..."):
        status = push_schedule_to_github(target_str)
        if status in [200, 201]: st.sidebar.success("✅ GitHub Updated!")
        else: st.sidebar.error(f"❌ Error: {status}")

# =====================================
# 3. Data Parsing Engine
# =====================================
def extract_and_flatten(df_raw):
    rows = []
    i = 0
    while i < len(df_raw):
        cell = str(df_raw.iloc[i, 0])
        if cell.startswith("Term:"):
            term = cell.replace("Term:", "").strip()
            header_row = i + 2
            headers = [str(h).strip() for h in df_raw.iloc[header_row].tolist()]
            j = header_row + 1
            while j < len(df_raw) and pd.notna(df_raw.iloc[j, 0]):
                row_data = dict(zip(headers, df_raw.iloc[j].tolist()))
                row_data["Term"] = term
                rows.append(row_data)
                j += 1
            i = j
        else: i += 1
    return pd.DataFrame(rows)

# =====================================
# 4. File Upload & State Management
# =====================================
uploaded_file = st.file_uploader("Upload Excel (.xlsx)", type=["xlsx"])

if uploaded_file:
    xls = pd.ExcelFile(uploaded_file)
    selected_sheet = st.selectbox("Select Subject", xls.sheet_names)
    state_key = f"df_{selected_sheet}"

    # Initialize Session State if new sheet or file
    if state_key not in st.session_state:
        df_raw = pd.read_excel(uploaded_file, sheet_name=selected_sheet, header=None)
        st.session_state[state_key] = extract_and_flatten(df_raw)

    # Use the Master Data from State
    master_df = st.session_state[state_key]

    # --- Term Selection Filter ---
    st.subheader("📅 Term Selection")
    all_terms = sorted(master_df["Term"].unique())
    select_terms_active = st.checkbox("Filter by specific terms")

    if select_terms_active:
        options = ["Select All"] + all_terms
        selected_terms = st.multiselect("Terms to process:", options, default=all_terms)
        if "Select All" in selected_terms: selected_terms = all_terms
    else:
        selected_terms = all_terms

    # Filtered Data for calculations
    df = master_df[master_df["Term"].isin(selected_terms)].copy()
    for s in skills: df[s] = pd.to_numeric(df[s], errors='coerce')

    # =====================================
    # 5. Data Editor (The Fix & Validation)
    # =====================================
    st.header(f"✏️ Data Review & Editor – {selected_sheet}")
    
    # Audit: Flagging Nulls
    audit_df = df.copy()
    audit_df.insert(0, "Status", audit_df.apply(lambda r: "🚨 MISSING" if r[skills].isnull().any() else "✅ OK", axis=1))
    
    show_nulls = st.checkbox("🔍 View only rows with 🚨")
    display_df = audit_df[audit_df["Status"] == "🚨 MISSING"] if show_nulls else audit_df

    edited_df = st.data_editor(display_df, num_rows="dynamic", use_container_width=True, key=f"editor_{state_key}")

    col_save, col_dl = st.columns(2)
    with col_save:
        if st.button("🔄 Apply Edits to Dashboard", use_container_width=True):
            cleaned_edits = edited_df.drop(columns=["Status"])
            st.session_state[state_key].update(cleaned_edits)
            st.success("Changes Saved to Session!")
            st.rerun()

    # 2. UPDATED FUNCTION
    def save_to_stacked_format(df_to_save, sheet_name, skill_cols):
        """Reconstructs the original Excel layout with Terms and dotted lines."""
        output = BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        current_row = 1
        
        # Sort terms using the global month_order
        terms = sorted(
            df_to_save["Term"].unique(), 
            key=lambda x: month_order.index(x) if x in month_order else 99
        )
        
        for term in terms:
            # Write Term Header
            ws.cell(row=current_row, column=1, value=f"Term: {term}").font = Font(bold=True)
            current_row += 1
            
            # Write Dotted Line
            ws.cell(row=current_row, column=1, value="--------------------------")
            current_row += 1
            
            # Write Table Headers
            headers = ["Student Name"] + skill_cols
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=current_row, column=col_num, value=header).font = Font(bold=True)
            current_row += 1
            
            # Write Student Data
            term_data = df_to_save[df_to_save["Term"] == term]
            for _, row in term_data.iterrows():
                ws.cell(row=current_row, column=1, value=row["Student Name"])
                for col_num, skill in enumerate(skill_cols, 2):
                    # We use .get() or index to ensure we match the right column
                    ws.cell(row=current_row, column=col_num, value=row[skill])
                current_row += 1
            
            # Add space between tables
            current_row += 2
            
        wb.save(output)
        return output.getvalue()
    
    # 3. UPDATED CALL (Inside your UI)
    with col_dl:
        # Pass the variables explicitly to avoid NameErrors
        stacked_excel_data = save_to_stacked_format(
            st.session_state[state_key], 
            selected_sheet, 
            skills
        )
        
        st.download_button(
            label="📥 Download Corrected Excel (Original Format)",
            data=stacked_excel_data,
            file_name=f"Corrected_{selected_sheet}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    # =====================================
    # 6. Calculations & Metrics
    # =====================================
    df["Average"] = df[skills].mean(axis=1)
    df["Grade"] = df["Average"].apply(lambda x: "A" if x>=80 else "B" if x>=70 else "C" if x>=60 else "D" if x>=50 else "F")
    df["Remarks"] = df["Average"].apply(
        lambda x: "Excellent work!" if x >= 80 else "Good effort!" if x >= 70 else "Needs improvement")
    
    st.divider()
    st.subheader("📌 Performance Highlights")
    m1, m2, m3 = st.columns(3)

    if not df.empty:
        # Metric 1: Top Student
        top_row = df.loc[df['Average'].idxmax()]
        m1.metric("🏆 Top Student", top_row['Student Name'], f"{top_row['Average']:.1f} Avg")

        # Metric 2: Growth Calculations
        df_melted = df.melt(id_vars=['Term'], value_vars=skills, var_name='Skill', value_name='Score')
        df_grouped = df_melted.groupby(['Term', 'Skill'])['Score'].mean().reset_index()
        
        if len(selected_terms) > 0:
            base_t = selected_terms[0]
            df_base = df_grouped[df_grouped['Term'] == base_t][['Skill', 'Score']].rename(columns={'Score':'Base'})
            df_final = pd.merge(df_grouped, df_base, on='Skill')
            df_final['Growth'] = ((df_final['Score'] - df_final['Base']) / df_final['Base']) * 100
            
            latest_g = df_final[df_final['Term'] == selected_terms[-1]]
            if not latest_g.empty:
                best_s = latest_g.loc[latest_g['Growth'].idxmax()]
                m2.metric("📈 Most Improved", best_s['Skill'], f"{best_s['Growth']:.1f}% Growth")

        # Metric 3: Success Rate
        rate = (len(df[df['Grade'].isin(['A', 'B'])]) / len(df)) * 100
        m3.metric("🎯 Class Success Rate", f"{rate:.0f}%", "Grades A & B")

        # =====================================
        # 6.5 STUDENT SEARCH FILTER (FIXED JOIN)
        # =====================================
        st.divider()
        st.subheader("🔍 Individual Student Search")
        
        student_names = sorted(df["Student Name"].dropna().unique().tolist())
        student_list = ["All Students"] + student_names
        search_query = st.selectbox("Search for a student:", student_list)
        
        if search_query != "All Students":
            active_df = df[df["Student Name"] == search_query].copy()
        else:
            active_df = df.copy()
        
        # --- DYNAMIC CALCULATION (Handles Missing Months) ---
        df_melted_active = active_df.melt(id_vars=['Term'], value_vars=skills, var_name='Skill', value_name='Score')
        df_final_active = df_melted_active.groupby(['Term', 'Skill'])['Score'].mean().reset_index()
        
        if not df_final_active.empty:
            # 1. Identify the earliest available month for the current selection
            # (Instead of forcing 'Jan', we take the first month they actually appear in)
            present_terms = [t for t in month_order if t in df_final_active['Term'].unique()]
            
            if present_terms:
                base_t = present_terms[0] 
                df_base = df_final_active[df_final_active['Term'] == base_t][['Skill', 'Score']].rename(columns={'Score':'Base'})
                
                # 2. Use 'how=left' to keep months like June/August even if Jan is missing
                df_final_active = pd.merge(df_final_active, df_base, on='Skill', how='left')
                
                # 3. Calculate Growth (Handle division by zero/NaN)
                df_final_active['Growth'] = ((df_final_active['Score'] - df_final_active['Base']) / df_final_active['Base']) * 100
                df_final_active['Growth'] = df_final_active['Growth'].fillna(0) # First month will show 0% growth
                
        # =====================================
        # 7, 8, & 9. SIDE-BY-SIDE ANALYTICS
        # =====================================
        st.header("📊 Subject Analytics & Insights")
        
        col_chart1, col_chart2 = st.columns(2)
        
        # Define your theme color to match the app background
        theme_bg = "#FFDE59" 

        # --- Left Column: Performance & Growth (Fused with Search) ---
        with col_chart1:
            st.subheader("Performance & Growth Trends")
            
            if not df_final_active.empty:
                base_chart = alt.Chart(df_final_active).encode(
                    x=alt.X('Term:N', sort=month_order, title="Academic Term")
                )
                
                # Bars for Scores
                bars = base_chart.mark_bar(opacity=0.4).encode(
                    xOffset='Skill:N',
                    y=alt.Y('Score:Q', scale=alt.Scale(domain=[0, 100]), title="Score"),
                    color=alt.Color('Skill:N', legend=alt.Legend(orient='bottom'))
                )
                
                # Line for Growth %
                lines = base_chart.mark_line(size=3, point=True).encode(
                    y=alt.Y('Growth:Q', title="Growth %", axis=alt.Axis(format='+')),
                    color='Skill:N',
                    tooltip=['Term', 'Skill', alt.Tooltip('Score:Q', format='.1f'), alt.Tooltip('Growth:Q', format='.1f')]
                )
                
                # Layering and applying the Yellow Background
                growth_chart = alt.layer(bars, lines).resolve_scale(y='independent').properties(
                    height=450,
                    background=theme_bg  # Matches your app color
                    ).configure_axis(
                    # Styling for the Titles (Score, Academic Term, Growth %)
                    titleColor='black',
                    titleFontSize=14,
                    titleFontWeight='bold',
                    
                    # Styling for the Labels (Jan, Feb, 0, 20, 40...)
                    labelColor='black',
                    labelFontSize=12,
                    labelFontWeight='bold',
                    
                    # Optional: Make the axis lines themselves black instead of grey
                    domainColor='black',
                    tickColor='black'
                    ).configure_legend(
                        titleColor='black',
                        labelColor='black',
                        labelFontWeight='bold',
                        titleFontWeight='bold'
                    ) #.configure_view(strokeOpacity=0)
                    
                st.altair_chart(growth_chart, use_container_width=True)
            else:
                st.warning("No data available for this selection.")
        
        # --- Right Column: Donut Chart (Top) & Statistics (Below) ---
        with col_chart2:
            st.subheader("Grade Distribution (%)")
            
            if not active_df.empty:
                base_pie = alt.Chart(active_df).encode(
                    theta=alt.Theta(field="Grade", aggregate="count", type="quantitative", stack=True),
                    color=alt.Color(
                        field="Grade", 
                        type="nominal", 
                        sort=grade_order, 
                        scale=alt.Scale(domain=grade_order, range=grade_colors),
                        legend=alt.Legend(title="Grades", orient="right")
                    )
                )
        
                pie = base_pie.mark_arc(innerRadius=60, outerRadius=140)
        
                text = base_pie.mark_text(radius=100, size=14, fontWeight="bold", color="white").encode(
                    text=alt.Text('pct:Q', format='.0%')
                ).transform_joinaggregate(
                    total='count(*)'
                ).transform_calculate(
                    pct='datum.count / datum.total'
                ).transform_filter(alt.datum.pct > 0.04)
        
                # Combining and applying the Yellow Background
                pie_chart = (pie + text).properties(
                    height=350,
                    background=theme_bg  # Matches your app color
                ).configure_legend(
                labelColor='black',
                titleColor='black',
                labelFontSize=12,
                labelFontWeight='bold',
                titleFontWeight='bold'
                )  #.configure_view(strokeOpacity=0)
        
                st.altair_chart(pie_chart, use_container_width=True)
        
                # 2. Statistic Reading (Below Donut)
                st.markdown("---")
                st.markdown("### 💡 **Analysis Insights**")
                
                avg_skills = active_df[skills].mean().sort_values()
                
                if not avg_skills.empty:
                    stat_col1, stat_col2 = st.columns(2)
                    with stat_col1:
                        st.markdown(f'''
                            <div style="background-color:#90EE90; padding:15px; border-radius:10px; border:2px solid black;">
                                <p style="margin:0; font-size:16px;">🌟 <b>Top Skill</b></p>
                                <p style="margin:0; font-size:20px;">{avg_skills.index[-1]}</p>
                                <p style="margin:0; font-size:14px;">Avg: {avg_skills.max():.1f}</p>
                            </div>
                        ''', unsafe_allow_html=True)
                    with stat_col2:
                        st.markdown(f'''
                            <div style="background-color:#D0312D; padding:15px; border-radius:10px; border:2px solid black; color:white;">
                                <p style="margin:0; font-size:16px; color:white;">⚠️ <b>Focus Area</b></p>
                                <p style="margin:0; font-size:20px; color:white;">{avg_skills.index[0]}</p>
                                <p style="margin:0; font-size:14px; color:white;">Avg: {avg_skills.min():.1f}</p>
                            </div>
                        ''', unsafe_allow_html=True)
                

    # =====================================
    # 8. Report Export & Google Drive
    # =====================================
    st.divider()
    col_zip, col_drive = st.columns(2)
    
    with col_zip:
        if st.button("📦 Generate Student PDF ZIP"):
            z_buf = BytesIO()
            # Use 'zf' as the zip handle
            with zipfile.ZipFile(z_buf, "w") as zf:
                for _, r in df.iterrows():
                    pdf = FPDF()
                    pdf.add_page()
                    
                    # --- 1. ADD LOGO ---
                    # Position: x=10, y=8 | Width: 33 (adjust as needed)
                    # Ensure 'logo.png' exists in your root folder
                    try:
                        pdf.image("logo.png", x=100, y=8, w=33)
                    except:
                        # Fallback if logo is missing to prevent crash
                        pdf.set_font("Arial", "I", 8)
                        pdf.cell(0, 5, "[School Logo Placeholder]", ln=True)
                    
                    # Move cursor down so text doesn't overlap logo
                    pdf.ln(20)
                    # Header
                    pdf.set_font("Arial", "B", 16)
                    # Use 'r' here to match your loop variable
                    pdf.cell(0, 10, f"Progress Report ({r['Term']})", ln=True)
                    
                    # Content
                    pdf.set_font("Arial", "", 12)
                    pdf.cell(0, 8, f"Student: {str(r['Student Name']).strip()}", ln=True)
                    pdf.cell(0, 5, "-"*30, ln=True) # Divider line
                    
                    # Skill Scores
                    for s in skills:
                        score = r[s] if pd.notna(r[s]) else 0
                        pdf.cell(0, 8, f"{s}: {score}", ln=True)
                    
                    # Final Stats
                    pdf.cell(0, 5, "-"*30, ln=True)
                    pdf.cell(0, 8, f"Average: {r['Average']:.2f}", ln=True)
                    pdf.cell(0, 8, f"Grade: {r['Grade']}", ln=True)
                    pdf.cell(0, 8, f"Remarks: {r['Remarks']}", ln=True)
                    
                    # Output PDF to string and write to ZIP
                    # FPDF output(dest="S") returns a string in latin-1
                    pdf_content = pdf.output(dest="S").encode("latin-1")
                    
                    # Organizing files in folders by Term
                    filename = f"{r['Term']}/{str(r['Student Name']).strip()}_report.pdf"
                    zf.writestr(filename, pdf_content)
            
            # Final download button
            st.download_button(
                label="⬇️ Download ZIP",
                data=z_buf.getvalue(),
                file_name=f"Student_Reports_{selected_sheet}.zip",
                mime="application/zip",
                use_container_width=True
            )

    with col_drive:
        folder_id_input = st.text_input("G-Drive Folder ID", "0ALncbMfl-gjdUk9PVA")
        if st.button("🚀 Upload Current View to Drive"):
            st.info("Initiating Google Drive Upload...")
            try:
                sa_info = json.loads(st.secrets["google_service_account"]["google_service_account"])
                credentials = service_account.Credentials.from_service_account_info(
                    sa_info, scopes=['https://www.googleapis.com/auth/drive']
                )
                drive_service = build('drive', 'v3', credentials=credentials)
        
                # --- FIX: Calculate EXACT total steps ---
                # Only count rows that belong to the selected terms
                df_to_process = df[df["Term"].isin(selected_terms)]
                total_steps = len(df_to_process)
                
                if total_steps == 0:
                    st.warning("No data found for the selected terms.")
                else:
                    prog_bar = st.progress(0)
                    status_text = st.empty() # Placeholder for "Processing Alice Tan..."
                    current_step = 0
        
                    for term in selected_terms:
                        term_clean = term.strip()
                        parent_id = folder_id_input.strip()
        
                        # 1. FIND OR CREATE TERM FOLDER
                        query_folder = (
                            f"name='{term_clean}' "
                            f"and mimeType='application/vnd.google-apps.folder' "
                            f"and '{parent_id}' in parents and trashed=false"
                        )
                        folder_search = drive_service.files().list(
                            q=query_folder, fields="files(id)", 
                            includeItemsFromAllDrives=True, supportsAllDrives=True
                        ).execute()
                        
                        existing_folders = folder_search.get('files', [])
                        if existing_folders:
                            term_folder_id = existing_folders[0]['id']
                        else:
                            folder_metadata = {'name': term_clean, 'mimeType': 'application/vnd.google-apps.folder', 'parents': [parent_id]}
                            term_folder = drive_service.files().create(body=folder_metadata, fields='id', supportsAllDrives=True).execute()
                            term_folder_id = term_folder['id']
        
                        # 2. UPLOAD / OVERWRITE PDFs
                        df_term = df[df["Term"] == term]
                        
                        for _, row in df_term.iterrows():
                            student_name = row['Student Name'].strip()
                            file_name = f"{selected_sheet}_{student_name}_report.pdf"
                            
                            # Update status text
                            status_text.text(f"Uploading: {file_name}")
        
                            # --- Generate PDF ---
                            pdf = FPDF()
                            pdf.add_page()

                            # Position: x=10, y=8 | Width: 33 (adjust as needed)
                            # Ensure 'logo.png' exists in your root folder
                            try:
                                pdf.image("logo.png", x=100, y=8, w=33)
                            except:
                                # Fallback if logo is missing to prevent crash
                                pdf.set_font("Arial", "I", 8)
                                pdf.cell(0, 5, "[School Logo Placeholder]", ln=True)
                            
                            # Move cursor down so text doesn't overlap logo
                            pdf.ln(20)
                            
                            pdf.set_font("Arial", "B", 16)
                            pdf.cell(0, 10, f"Progress Report ({selected_sheet}_{term_clean})", ln=True)
                            pdf.set_font("Arial", "", 12)
                            pdf.cell(0, 8, f"Student: {student_name}", ln=True)
                            pdf.cell(0, 5, "-"*30, ln=True) # Divider line
                            for s in skills:
                                pdf.cell(0, 8, f"{s}: {row[s]}", ln=True)
                            pdf.cell(0, 5, "-"*30, ln=True)
                            pdf.cell(0, 8, f"Average: {row['Average']:.2f}", ln=True)
                            pdf.cell(0, 8, f"Grade: {row['Grade']}", ln=True)
                            pdf.cell(0, 8, f"Remarks: {row['Remarks']}", ln=True)
        
                            pdf_bytes = BytesIO()
                            pdf_bytes.write(pdf.output(dest="S").encode("latin-1"))
                            pdf_bytes.seek(0)
                            media = MediaIoBaseUpload(pdf_bytes, mimetype='application/pdf', resumable=True)
        
                            # --- Check if FILE already exists for Overwrite ---
                            query_file = f"name='{file_name}' and '{term_folder_id}' in parents and trashed=false"
                            file_search = drive_service.files().list(q=query_file, fields="files(id)", includeItemsFromAllDrives=True, supportsAllDrives=True).execute()
                            existing_files = file_search.get('files', [])
        
                            if existing_files:
                                drive_service.files().update(fileId=existing_files[0]['id'], media_body=media, supportsAllDrives=True).execute()
                            else:
                                file_metadata = {'name': file_name, 'parents': [term_folder_id]}
                                drive_service.files().create(body=file_metadata, media_body=media, fields='id', supportsAllDrives=True).execute()
        
                            # --- FIX: Increment progress based on ACTUAL processed rows ---
                            current_step += 1
                            prog_bar.progress(current_step / total_steps)
        
                    status_text.empty() # Clear status when done
                    st.success(f"✅ Successfully processed {total_steps} reports!")
        
            except Exception as e:
                st.error(f"Google Drive operation failed: {e}")
        

